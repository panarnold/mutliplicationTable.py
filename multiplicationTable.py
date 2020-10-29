#! python
# multiplicationTable.py - script responsible for creating multiplication table of number which you give it
# in sys.args
# X 2020 Arnold Cytrowski


import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv)!=2:
	print('Usage: python3 multiplicationTable.py <int>')
	sys.exit(-1)

number = None

try:
	number = abs(int(sys.argv[1]))
except ValueError as e:
	print('First argument should be an integer')
	sys.exit(-1)

bold_font = Font(bold=True)


wb = openpyxl.Workbook()
sheet = wb.active
for row_num in range(1, number + 2):
    for column_num in range(1, number + 2):
        if row_num == 1 and column_num == 1:
            sheet.cell(row = 1, column = 1).value='*'
        elif row_num == 1:
            sheet.cell(row = row_num, column = column_num).value = column_num - 1
            sheet.cell(row = row_num, column = column_num).font = bold_font
        elif column_num == 1:
            sheet.cell(row = row_num, column = column_num).value = row_num - 1
            sheet.cell(row = row_num, column = column_num).font = bold_font
            
        else:
            sheet.cell(row = row_num, column = column_num).value = (row_num-1)*(column_num - 1)



wb.save(f'multiplicationTablefor{number}number.xlsx')
wb.close()
sys.exit(0)

